# styling the worksheet in openpyxl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Defining custom styles
font1 = Font(name='Arial', size=12, bold=True)
alignment = Alignment(horizontal='center', vertical='center')
rfill = PatternFill(patternType='solid', fgColor='D492a0')


# Function to style the workbook saved by calculator function
# Takes the path to that workbook as the argument
def style(workbook_path):
    letters = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L') #the columns
    sizes = (15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 20, 20) #column width sizes

    wb = openpyxl.load_workbook(workbook_path) #load the workbook
    ws = wb['Sheet1'] # reading the required worksheet
    ws.row_dimensions[1].height = 30 #height for row 1 (with headings)
    ws.row_dimensions[2].height = 15 #height for row 2
    ws.insert_rows(2, 1) #insert a row below the headings

    indices = range(len(letters)) #looping range columnwise

    # Set each column with the corrsponding width
    for i in indices:
        ws.column_dimensions[letters[i]].width = sizes[i]

    # Set font and alignment for each below the heading row
    for j in range(1, 13):
        ws.cell(row=1, column=j).font = font1
        ws.cell(row=1, column=j).alignment = alignment

    ws.freeze_panes = 'A3' #freeze the top two rows

    # Merge top two rows in columns A, B, C, D, E
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')

    #Add extra information on the limitations in columns computed by calculator
    ws['F2'].value = '(Max 10hrs)'
    ws['G2'].value = '(Max 160hrs)'
    ws['H2'].value = '(Max 100hrs)'
    ws['I2'].value = '(Max 1000hrs)'
    ws['J2'].value = '(Min 12hrs)'
    ws['K2'].value = '(Min 2 days)'
    ws['L2'].value = '(Min 8 days)'


    # Get list of all rows and the location as an index (prevent zero indexing)
    # Set the height for all rows starting from 3 to end
    # Frozen rows excluded, their size already set above
    rows = list(ws.rows)
    row_index = range(3, len(rows) + 1)

    for k in row_index:
        ws.row_dimensions[k].height = 20

    # Setting alignment for each cell
    for m in range(1, 13): #12 columns without zero indexing
        for n in range(2, len(rows) + 1): #each row below the top
            ws.cell(row=n, column=m).alignment = alignment

    max_limits = [10, 160, 100, 1000]
    min_limits = [12, 2, 8]
    i = 0

    # Different style for cells whose values are outside acceptable limits
    # This is important to highlight the violations

    for col in range(6, 10): #the calculated columns
        for row in range(3, len(rows) + 1): #all unfrozen cells (non headers)
            if ws.cell(row=row, column=col).value > max_limits[i]:
                ws.cell(row=row, column=col).fill = rfill
                ws.cell(row=row, column=col).font = font1
        i += 1

    i = 0
    for col in range(10, 13):
        for row in range(3, len(rows) + 1):
            if ws.cell(row=row, column=col).value < min_limits[i]:
                ws.cell(row=row, column=col).fill = rfill
                ws.cell(row=row, column=col).font = font1
        i += 1

    wb.save(workbook_path)
